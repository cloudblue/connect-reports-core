#  Copyright © 2021 CloudBlue. All rights reserved.

import json
import os
from datetime import datetime

import pytz

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color, WHITE
from openpyxl.utils.exceptions import InvalidFileException

from zipfile import BadZipfile

from connect.reports.renderers.base import BaseRenderer
from connect.reports.renderers.registry import register


@register('xlsx')
class XLSXRenderer(BaseRenderer):
    def render(self, data, output_file, start_time=None):
        self.start_time = start_time or datetime.now(tz=pytz.utc)
        return self.generate_report(data, output_file)

    def generate_report(self, data, output_file):
        start_col_idx = self.args.get('start_col', 1)
        row_idx = self.args.get('start_row', 2)
        wb = load_workbook(
            os.path.join(
                self.root_dir,
                self.template,
            ),
        )
        ws = wb['Data']
        for row in data:
            for col_idx, cell_value in enumerate(row, start=start_col_idx):
                ws.cell(row_idx, col_idx, value=cell_value)
            row_idx += 1

        self._add_info_sheet(wb.create_sheet('Info'), self.start_time)

        output_file = f'{output_file}.xlsx'
        wb.save(output_file)
        return output_file

    def _add_info_sheet(self, ws, start_time):
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 180
        ws['A1'].value = 'Report Execution Information'
        ws.merge_cells('A1:B1')
        ws['A1'].fill = PatternFill('solid', start_color=Color('1565C0'))
        ws['A1'].font = Font(sz=24, color=WHITE)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        for i in range(2, 10):
            ws[f'A{i}'].alignment = Alignment(
                horizontal='left',
                vertical='top',
            )
            ws[f'B{i}'].alignment = Alignment(
                horizontal='left',
                vertical='top',
            )
        ws['A2'].value = 'Report Start time'
        ws['B2'].value = start_time.strftime('%Y-%m-%d %H:%M:%S')
        ws['A3'].value = 'Report Finish time'
        ws['B3'].value = datetime.now(tz=pytz.utc).strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'].value = 'Account ID'
        ws['B4'].value = self.account.id
        ws['A5'].value = 'Account Name'
        ws['B5'].value = self.account.name
        ws['A6'].value = 'Report ID'
        ws['B6'].value = self.report.id
        ws['A7'].value = 'Report Name'
        ws['B7'].value = self.report.name
        ws['A8'].value = 'Runtime environment'
        ws['B8'].value = self.environment
        ws['A9'].value = 'Report execution parameters'
        ws['B9'].value = json.dumps(self.report.values, indent=4, sort_keys=True)
        ws['B9'].alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True,
        )

    @classmethod
    def _validate_args(cls, args):
        errors = []
        start_row = args.get('start_row')
        start_col = args.get('start_col')
        if start_row is not None:
            if not isinstance(start_row, int):
                errors.append('`start_row` must be integer.')
            else:
                if start_row < 1:
                    errors.append('`start_row` must be greater than 0.')
        if start_col is not None:
            if not isinstance(start_col, int):
                errors.append('`start_col` must be integer.')
            else:
                if start_col < 1:
                    errors.append('`start_col` must be greater than 0.')
        return errors

    @classmethod
    def validate(cls, definition):
        errors = []
        if definition.template is None:
            errors.append('`template` is required for xlsx renderer.')
            return errors
        elif not os.path.isfile(
            os.path.join(definition.root_path, definition.template),
        ):
            errors.append(f'template `{definition.template}` not found.')
            return errors

        try:
            template_file = os.path.join(definition.root_path, definition.template)
            load_workbook(template_file)
        except (InvalidFileException, BadZipfile):
            errors.append(f'template `{definition.template}` not valid or empty.')

        if definition.args is not None:
            errors.extend(cls._validate_args(definition.args))
        return errors
