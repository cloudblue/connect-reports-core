#  Copyright © 2022 CloudBlue. All rights reserved.

from datetime import datetime

import pytest

from fs.tempfs import TempFS

from openpyxl import Workbook, load_workbook

from connect.reports.datamodels import RendererDefinition
from connect.reports.renderers import XLSXRenderer


@pytest.mark.parametrize('args', (None, {}, {'start_row': 1, 'start_col': 1}))
def test_validate_ok(mocker, args):
    tmp_fs = TempFS()
    defs = RendererDefinition(
        root_path=tmp_fs.root_path,
        id='renderer_id',
        type='xlsx',
        description='description',
        template='template.xlsx',
        args=args,
    )
    _create_xlsx_doc(f'{tmp_fs.root_path}/{defs.template}')

    assert XLSXRenderer.validate(defs) == []


def test_validate_no_template():
    defs = RendererDefinition(
        root_path='root_path',
        id='renderer_id',
        type='xlsx',
        description='description',
    )

    assert XLSXRenderer.validate(defs) == ['`template` is required for xlsx renderer.']


def test_validate_template_not_found(mocker):
    mocker.patch('connect.reports.renderers.xlsx.os.path.isfile', return_value=False)
    defs = RendererDefinition(
        root_path='root_path',
        id='renderer_id',
        type='xlsx',
        description='description',
        template='template.xlsx',
    )

    assert XLSXRenderer.validate(defs) == ['template `template.xlsx` not found.']


@pytest.mark.parametrize(
    ('args', 'error'),
    (
        ({'start_row': 'a'}, '`start_row` must be integer.'),
        ({'start_col': 'a'}, '`start_col` must be integer.'),
        ({'start_row': 0}, '`start_row` must be greater than 0.'),
        ({'start_row': -3}, '`start_row` must be greater than 0.'),
        ({'start_col': 0}, '`start_col` must be greater than 0.'),
        ({'start_col': -3}, '`start_col` must be greater than 0.'),
    ),
)
def test_validate_invalid_args(mocker, args, error):
    tmp_fs = TempFS()
    defs = RendererDefinition(
        root_path=tmp_fs.root_path,
        id='renderer_id',
        type='xlsx',
        description='description',
        template='template.xlsx',
        args=args,
    )
    _create_xlsx_doc(f'{tmp_fs.root_path}/{defs.template}')

    assert XLSXRenderer.validate(defs) == [error]


def test_generate_report(mocker, account_factory, report_factory, report_data):
    data = report_data()

    expected_calls = []
    for i in range(10):
        for j in range(10):
            expected_calls.append(mocker.call(2 + i, 1 + j, value=f'row_{i}_col_{j}'))

    data_sheet = mocker.MagicMock()

    wbmock = mocker.MagicMock()
    wbmock.__getitem__.side_effect = [data_sheet]

    mocked_load_wb = mocker.patch(
        'connect.reports.renderers.xlsx.load_workbook',
        return_value=wbmock,
    )
    acc = account_factory()
    report = report_factory()

    renderer = XLSXRenderer(
        'runtime environment', 'root_path', acc, report, template='template.xlsx',
    )

    renderer.generate_summary = mocker.MagicMock()
    renderer.start_time = datetime.utcnow()

    assert renderer.generate_report(data, 'report') == 'report.xlsx'
    mocked_load_wb.assert_called_once_with('root_path/template.xlsx')
    data_sheet.cell.assert_has_calls(expected_calls)
    wbmock.save.assert_called_once_with('report.xlsx')


def test_validate_template_not_valid():

    tmp_filesystem = TempFS()
    tmp_filesystem.create('test.xlsx')

    defs = RendererDefinition(
        root_path=tmp_filesystem.root_path,
        id='renderer_id',
        type='xlsx',
        description='description',
        template='test.xlsx',
    )

    errors = XLSXRenderer.validate(defs)

    assert 'not valid or empty' in errors[0]


def test_render_tmpfs_ok(account_factory, report_factory, report_data):
    tmp_fs = TempFS()
    tmp_fs.makedirs('package/report')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.cell(1, 1, value='Name')
    ws.cell(1, 2, value='Description')
    wb.save(f'{tmp_fs.root_path}/package/report/template.xlsx')

    renderer = XLSXRenderer(
        'runtime',
        tmp_fs.root_path,
        account_factory(),
        report_factory(),
        template='package/report/template.xlsx',
    )

    data = report_data(2, 2)
    path_to_output = f'{tmp_fs.root_path}/package/report/report'
    output_file = renderer.render(data, path_to_output, start_time=datetime.now())

    wb = load_workbook(output_file)
    ws = wb['Data']

    assert output_file == f'{path_to_output}.xlsx'
    assert data == [[ws[f'A{item}'].value, ws[f'B{item}'].value] for item in range(2, 4)]


@pytest.mark.asyncio
async def test_render_async_tmpfs_ok(account_factory, report_factory, report_data):
    tmp_fs = TempFS()
    tmp_fs.makedirs('package/report')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.cell(1, 1, value='Name')
    ws.cell(1, 2, value='Description')
    wb.save(f'{tmp_fs.root_path}/package/report/template.xlsx')

    renderer = XLSXRenderer(
        'runtime',
        tmp_fs.root_path,
        account_factory(),
        report_factory(),
        template='package/report/template.xlsx',
    )

    data = report_data(2, 2)

    async def async_gen(cols, rows):
        for element in report_data(cols, rows):
            yield element

    path_to_output = f'{tmp_fs.root_path}/package/report/report'
    output_file = await renderer.render_async(
        async_gen(2, 2),
        path_to_output,
        start_time=datetime.now(),
    )
    wb = load_workbook(output_file)
    ws = wb['Data']

    assert output_file == f'{path_to_output}.xlsx'
    assert data == [[ws[f'A{item}'].value, ws[f'B{item}'].value] for item in range(2, 4)]


@pytest.mark.asyncio
async def test_render_async_with_sync_data(account_factory, report_factory, report_data):
    tmp_fs = TempFS()
    tmp_fs.makedirs('package/report')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.cell(1, 1, value='Name')
    ws.cell(1, 2, value='Description')
    wb.save(f'{tmp_fs.root_path}/package/report/template.xlsx')

    renderer = XLSXRenderer(
        'runtime',
        tmp_fs.root_path,
        account_factory(),
        report_factory(),
        template='package/report/template.xlsx',
    )

    data = report_data(2, 2)

    def sync_gen(cols, rows):
        for element in report_data(cols, rows):
            yield element

    path_to_output = f'{tmp_fs.root_path}/package/report/report'
    output_file = await renderer.render_async(
        sync_gen(2, 2),
        path_to_output,
        start_time=datetime.now(),
    )
    wb = load_workbook(output_file)
    ws = wb['Data']

    assert output_file == f'{path_to_output}.xlsx'
    assert data == [[ws[f'A{item}'].value, ws[f'B{item}'].value] for item in range(2, 4)]


def _create_xlsx_doc(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    for _ in range(1, 10):
        ws.append(range(10))
    wb.save(xlsx_path)
